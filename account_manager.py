import openpyxl as op
from openpyxl import load_workbook
from openpyxl import styles as st
from openpyxl.styles import PatternFill
import pandas as pd
import datetime 
import tkinter as tk
from tkinter import filedialog , simpledialog

class Account:
    def __init__(self) -> None:
        # Load the workbook and select the active sheet
        root=tk.Tk()
        root.withdraw()
        self.file = filedialog.askopenfilename(title="Select a File", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        print("File path",self.file)
        self.wb = load_workbook(rf"{self.file}", data_only=True)
        self.sheet = self.wb.active
        self.rawData = pd.read_excel(rf"{self.file}")
        # Initialize an empty self.dataFrame
        self.data = pd.DataFrame({
            "Data Type": [],
            "Total Debit": [],
            "Total Credit": [],
        })
        
        # Store all row data in a list of dictionaries
        self.row_data = []  # To store all row data
        
        for raw in range(2, self.sheet.max_row + 1):
            col_date = self.sheet[f"A{raw}"].value
            col_name = self.sheet[f"B{raw}"].value
            col_for = self.sheet[f"C{raw}"].value
            col_debit = self.sheet[f"D{raw}"].value
            col_credit = self.sheet[f"E{raw}"].value
            
            if col_name == "" or col_name == None:
                self.sheet.delete_rows(raw)
            # Store each row as a dictionary
            self.row_data.append({
                "date": col_date,
                "Name": col_name,
                "for": col_for,
                "debit": col_debit if col_debit is not None else 0,
                "credit": col_credit if col_credit is not None else 0,
            })
            
    # To update data
    def update_data(self):
        file=self.sheet
        row = file.max_row+1
        update_times = int(input("How many lines of updates you want to made : "))
        #delete last total 
        if file[f"C{row-1}"].value == "Total":
            file.delete_rows(row-1)
        #delete if any row is empty
        for i in range(2,row):
            if file[f"B{i}"] == "" or file[f"B{i}"] == None:
                file.delete_rows(i)
        #create new rows and take input from user to add data
        for i in range(update_times,0,-1):
            file.insert_rows(row)
            balance_row = file.max_row-1
            file.insert_rows(row)
            print(f"Row {update_times - i}")
            date = input("Enter Date DD-MM-YYYY: ")
            date = datetime.datetime.strptime(date , "%d-%m-%Y")
            date = date.strftime("%d-%m-%Y")
            name = input("Enter Name : ").upper()
            col_for = input("Eneter Data Type : ").upper()
            debit = (input("Enter Debit(dr.) amount : "))
            credit = (input("Enter Cerdit(cr.) amount : "))
            
            debit = float(debit) if debit and debit.strip() else 0
            credit = float(credit) if credit and credit.strip() else 0

            #specify data is debit or credit
            if debit == 0 :
                name = "BY " + name 
            else:
                name = "TO " + name
            #calculate remaning balance
            balance_value = file[f"F{balance_row}"].value
            balance_value_int=float(balance_value) if balance_value and str(balance_value).strip() else 0
            balance = (balance_value_int - debit) + credit

            file[f"A{row}"] = date
            file[f"A{row}"].number_format = "DDMMYY"
            file[f"B{row}"] = name
            file[f"B{row}"].number_format = st.numbers.FORMAT_GENERAL 
            file[f"C{row}"] = col_for
            file[f"C{row}"].number_format = st.numbers.FORMAT_GENERAL 
            file[f"D{row}"] = debit
            file[f"D{row}"].number_format = '₹#,##0.00'
            file[f"E{row}"] = credit
            file[f"E{row}"].number_format = '₹#,##0.00'
            file[f"F{row}"] = balance
            file[f"F{row}"].number_format = '₹#,##0.00'
            self.wb.save(self.file)
        # add total row in last 
        total_debit= 0
        total_credit=0
        for i in range(2,file.max_row+1):
            total_dr = file[f"D{i}"].value
            total_debit += float(total_dr) if total_dr and str(total_dr).strip() else 0
            total_cr = file[f"E{i}"].value
            total_credit += float(total_cr) if total_cr and str(total_cr).strip() else 0

        last_row=["","","Total",total_debit,total_credit,(total_credit-total_debit)]
        file.append(last_row)
        self.wb.save(self.file)
        #color total row
        yellow_color = PatternFill(start_color="FFFF00", fill_type="solid")
        orange_color = PatternFill(start_color="E46D0A", fill_type="solid")
        max_row_af_save = self.sheet.max_row
        file[f"C{max_row_af_save}"].fill = yellow_color
        file[f"D{max_row_af_save}"].fill = yellow_color
        file[f"E{max_row_af_save}"].fill = yellow_color
        file[f"F{max_row_af_save}"].fill = orange_color
        self.wb.save(self.file)
        
        ask=input("Do you want to made more updates Y/N : ").upper()
        if ask == "Y":
            return self.update_data()
        elif ask == "" or ask =="N":
            return "Data Updated successfully."

    # Calculate Total balance of Total debit and Total Credit according to data type
    def balance(self):
        TotalD = 0
        TotalC = 0
        
        # Loop through the stored rows
        for row in self.row_data:
            col_name = row['for']
            
            if col_name == None:
                for i in self.data.iterrows():
                    TotalD += row["Total Debit"] if "Total Debit" in row and pd.notna(row["Total Debit"]) else 0
                    TotalC += row["Total Credit"] if "Total Credit" in row and pd.notna(row["Total Credit"]) else 0
                self.data = self.data.append({
                    "Data Type": "Total Balance",
                    "Total Debit": TotalD,
                    "Total Credit": TotalC
                }, ignore_index=True)
                break
            elif col_name == "TOTAL":
                pass
            # Append the row data to the DataFrame
            elif not self.data["Data Type"].isin([col_name]).any():
                self.data = self.data.append({
                    "Data Type": col_name,
                    "Total Debit": row["debit"],
                    "Total Credit": row["credit"]
                }, ignore_index=True)
            else:
                for i in range(len(self.data)):
                    if self.data["Data Type"][i] == col_name:
                        if row["debit"] != 0:
                            self.data.loc[i, "Total Debit"] += row["debit"]
                        else:
                            self.data.loc[i, "Total Credit"] += row["credit"]
        return self.data

    def check_info_by_data(self):
        user = input("Enter Name : ").upper()
        useFor = input("Enter Data Type : ").upper() 
        date=input("Search According Date DD-MM-YYYY: ")
        total_d = []
        total_c = []
        data = pd.DataFrame({
            "Date": [],
            "Particular": [],
            "Data Type": [],
            "Debit": [],
            "Credit": []
        })
        no_transaction = 0
        
        if date:
            try:
                date_obj = datetime.datetime.strptime(date , "%d-%m-%Y").date()
            except ValueError:
                print("Invalid date format. Please use DD-MM-YYYY format.")
                return
        else:
            date_obj = None
        for row in self.row_data:
            col_name = row["Name"]
            col_for = row["for"]
            col_date=row["date"]

            if isinstance(col_date, str):
                try:
                    col_date = datetime.datetime.strptime(col_date, "%Y-%m-%d %H:%m:%s").date()
                    col_date = col_date.strftime("%d-%m-%Y")
                except ValueError:
                    print(f"Invalid date formate in data: {col_date}")
                    continue  # Skip rows with invalid date format
            elif isinstance(col_date, datetime.datetime):
                col_date=col_date.date()
                
            if col_name is not None:
                col_name = col_name[3:]
                 # Trimming part of the name if needed
                # Convert col_date to a date object if it's a string

                if (user == col_name and useFor == col_for and date_obj == col_date)\
                    or (user == col_name and useFor == col_for and not date_obj)\
                    or (user == col_name and not useFor and date_obj == col_date)\
                    or (user == col_name and not date_obj and not useFor)\
                    or (not user and useFor == col_for and not date_obj)\
                    or (not user and useFor == col_for and date_obj == col_date)\
                    or (not user and not useFor and date_obj == col_date):
                    no_transaction += 1
                    data = data.append({
                        "Date": row["date"],
                        "Particular": col_name,
                        "Data Type": col_for,
                        "Debit": row["debit"],
                        "Credit": row["credit"],
                    }, ignore_index=True)

                    if row["debit"] and row["credit"]:
                        total_d.append(row["debit"])
                        total_c.append(row["credit"])
                    elif row["debit"] != 0.0:
                        total_d.append(row["debit"])
                    else:
                        total_c.append(row["credit"])
            elif(not user and not useFor and not date):
                    return Account.balance(self)

        Total_Dr=0
        Total_Cr=0
        for i in total_d:
            if i != 0:
                Total_Dr+=i
        for i in total_c:
            if i != 0: 
                Total_Cr+=i

        no_transaction=(f"Total Number of Transactions: {no_transaction}")
        ST1=(f"Total Debit  = {Total_Dr}")
        ST2=(f"Total Credit = {Total_Cr}")
        return (f"{data}\n{ST1}\n{ST2}\n{no_transaction}")

class NewAccount(Account):
    def new_file(self):
        root = tk.Tk()
        root.withdraw()
        # folder_path = simpledialog.askstring(title="Select Folder")
        # if not folder_path:
        #     print("No Folder selected")
        #     return None  # Stop here if no folder is selected

        file_name = simpledialog.askstring("Input", "Enter file name")
        if not file_name:
            print("No file name entered")
            return None  # Stop here if no filename is entered
        wb = op.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        title_col =  PatternFill(start_color="B8CCE4",fill_type="solid")
        header = ["Date","Name","Data Type","Debit","Credits","Balance"]
        for col_num,head in enumerate(header, start=1):
            cell=ws.cell(row=1,column=col_num,value=head)
            cell.fill=title_col

        wb.save(self.file_path)
        
        return self.file_path 

    def new_file_rows(self):
        file_path=(NewAccount.new_file(self))
        wb=load_workbook(file_path)
        sheet=wb.active

        print("Add Your First Entry")
        date=datetime.datetime.now()
        date=(date.strftime("%d-%m-%Y"))
        particular=input("Enter Name : ").upper()
        data_type=input("Data Type : ").upper()
        Debit=float(input("Debit Amount : ")or 0)
        credit=float(input("Credit Amount : ")or 0)

        if Debit:
            particular = "TO " + particular
            credit = 0
        else:
            particular = "BY " + particular
            Debit = 0

        balance=credit-Debit

        first_row=[date,particular,data_type,Debit,credit,balance]
        sheet.append(first_row)
        print(f"Date : {date}")
        print(f"Name : {particular}")
        print(f"Data Type : {data_type}")
        print(f"Debit Amount : {Debit}")
        print(f"Credit Amount : {credit}")
        print(f"Balance : {balance}")

        wb.save(file_path)
        ask= input("Want Enter More Entrys Y/N : ").upper()
        if ask == "Y":
            Account.update_data(file_path)
        else:
            return "New File created Successfully."
# create an instance of Account and use the check_info_by_data function
if __name__ == "__main__":
    user_want=input("Want to create new file Y/N : ").upper()
    if user_want == "Y":
        file2 = NewAccount()
        f2 = file2.new_file_rows()
    else:
        # making choice
        file1 = Account()
        func_list=["Update Data","Check Balance", "Check By Info"]
        for i,option in enumerate(func_list, start=1): 
            print(f"{i}: {option}.")

        choice=int(input("Enter the number or name of the option: "))
        choices = {1: file1.update_data, 2: file1.balance, 3: file1.check_info_by_data}
        ch=choices.get(choice, lambda: print("No Choice has been made."))()
        print(ch)